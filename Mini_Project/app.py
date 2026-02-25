"""
============================================================
 3-PHASE POWER FACTOR MONITOR — Python Flask Application
============================================================
 Receives 3-phase sensor data wirelessly from ESP32 via
 HTTP POST, logs to Excel, sends desktop notifications,
 and serves a real-time web dashboard.

 Usage:
   python app.py                    # Normal mode
   python app.py --simulate         # Simulation mode (fake 3-phase data)
   python app.py --port 5000        # Custom port
   python app.py --threshold 0.85   # Custom PF threshold
============================================================
"""

import os
import sys
import json
import time
import random
import math
import argparse
import threading
from datetime import datetime
from collections import deque

# Fix Windows console encoding (cp1252 can't render emoji)
if sys.platform == 'win32':
    sys.stdout.reconfigure(errors='replace')
    sys.stderr.reconfigure(errors='replace')

from flask import Flask, request, jsonify, render_template
from openpyxl import Workbook, load_workbook

# ── Try to import plyer for notifications ────────────────────
try:
    from plyer import notification
    NOTIFICATIONS_AVAILABLE = True
except ImportError:
    NOTIFICATIONS_AVAILABLE = False
    print("⚠  plyer not installed — desktop notifications disabled.")
    print("   Install with: pip install plyer")

# ════════════════════════════════════════════════════════════
#  CONFIGURATION
# ════════════════════════════════════════════════════════════
DEFAULT_PORT = 5000
DEFAULT_THRESHOLD = 0.85
EXCEL_FILE = "power_factor_log.xlsx"
MAX_READINGS = 500
NOTIFICATION_COOLDOWN = 30
PHASE_NAMES = ['R', 'Y', 'B']
PHASE_KEYS = ['phase_r', 'phase_y', 'phase_b']

# ════════════════════════════════════════════════════════════
#  FLASK APP
# ════════════════════════════════════════════════════════════
app = Flask(__name__)

readings = deque(maxlen=MAX_READINGS)
last_notification_time = 0
pf_threshold = DEFAULT_THRESHOLD

excel_lock = threading.Lock()


# ════════════════════════════════════════════════════════════
#  EXCEL LOGGER — 3-Phase
# ════════════════════════════════════════════════════════════
def init_excel():
    """Create Excel file with 3-phase headers."""
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "3-Phase Power Factor Log"

        headers = ["Timestamp"]
        for phase in PHASE_NAMES:
            headers += [
                f"V_{phase} (V)",
                f"I_{phase} (A)",
                f"PF_{phase}",
                f"P_{phase} (W)",
                f"S_{phase} (VA)",
                f"Q_{phase} (VAR)",
            ]
        headers += [
            "Overall PF",
            "Total P (W)",
            "Total S (VA)",
            "Total Q (VAR)",
            "Status"
        ]
        ws.append(headers)

        # Style headers
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # Phase colors: R=Red, Y=Yellow, B=Blue
        phase_colors = ['D32F2F', 'F9A825', '1565C0']
        general_color = '2E7D32'

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = thin_border

            # Choose header color by phase
            if col_num == 1:
                color = '37474F'
            elif col_num <= 7:
                color = phase_colors[0]
            elif col_num <= 13:
                color = phase_colors[1]
            elif col_num <= 19:
                color = phase_colors[2]
            else:
                color = general_color
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

        # Column widths
        for i in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = 13
        ws.column_dimensions['A'].width = 20

        wb.save(EXCEL_FILE)
        print(f"📊 Created Excel file: {EXCEL_FILE}")


def log_to_excel(data):
    """Append a 3-phase reading to Excel."""
    with excel_lock:
        try:
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active

            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

            overall_pf = data.get('overall_pf', 0)
            status = "✅ Good" if overall_pf >= pf_threshold else "⚠️ Low PF"

            row = [data['timestamp']]
            for key in PHASE_KEYS:
                phase = data.get(key, {})
                row += [
                    round(phase.get('voltage', 0), 2),
                    round(phase.get('current', 0), 3),
                    round(phase.get('power_factor', 0), 3),
                    round(phase.get('real_power', 0), 2),
                    round(phase.get('apparent_power', 0), 2),
                    round(phase.get('reactive_power', 0), 2),
                ]
            row += [
                round(overall_pf, 3),
                round(data.get('total_real_power', 0), 2),
                round(data.get('total_apparent_power', 0), 2),
                round(data.get('total_reactive_power', 0), 2),
                status
            ]

            ws.append(row)

            last_row = ws.max_row
            num_cols = len(row)
            for col in range(1, num_cols + 1):
                cell = ws.cell(row=last_row, column=col)
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border

            # Color status cell
            status_cell = ws.cell(row=last_row, column=num_cols)
            if overall_pf < pf_threshold:
                status_cell.fill = PatternFill(start_color='FFD5D5', end_color='FFD5D5', fill_type='solid')
                status_cell.font = Font(color='CC0000', bold=True)
            else:
                status_cell.fill = PatternFill(start_color='D5FFD5', end_color='D5FFD5', fill_type='solid')
                status_cell.font = Font(color='006600', bold=True)

            # Also highlight individual low-PF phase cells
            for i, key in enumerate(PHASE_KEYS):
                pf_col = 2 + i * 6 + 2  # PF column for each phase
                phase_pf = data.get(key, {}).get('power_factor', 0)
                pf_cell = ws.cell(row=last_row, column=pf_col)
                if phase_pf > 0.01 and phase_pf < pf_threshold:
                    pf_cell.font = Font(color='CC0000', bold=True)

            wb.save(EXCEL_FILE)
        except Exception as e:
            print(f"❌ Excel write error: {e}")


# ════════════════════════════════════════════════════════════
#  DESKTOP NOTIFICATIONS
# ════════════════════════════════════════════════════════════
def send_notification(data):
    """Send notification if any phase PF is below threshold."""
    global last_notification_time

    if not NOTIFICATIONS_AVAILABLE:
        return

    # Check which phases have low PF
    low_phases = []
    for i, key in enumerate(PHASE_KEYS):
        pf = data.get(key, {}).get('power_factor', 1.0)
        if 0.01 < pf < pf_threshold:
            low_phases.append(f"Phase {PHASE_NAMES[i]}: {pf:.3f}")

    overall_pf = data.get('overall_pf', 1.0)
    if overall_pf >= pf_threshold and not low_phases:
        return

    now = time.time()
    if now - last_notification_time < NOTIFICATION_COOLDOWN:
        return
    last_notification_time = now

    try:
        msg = f"Overall PF: {overall_pf:.3f} (Threshold: {pf_threshold})\n"
        if low_phases:
            msg += "Low phases:\n" + "\n".join(low_phases)

        notification.notify(
            title="⚡ 3-Phase Power Factor Alert!",
            message=msg,
            app_name="Power Factor Monitor",
            timeout=10
        )
        print("🔔 Desktop notification sent!")
    except Exception as e:
        print(f"⚠  Notification error: {e}")


# ════════════════════════════════════════════════════════════
#  FLASK ROUTES
# ════════════════════════════════════════════════════════════

@app.route('/')
def dashboard():
    return render_template('index.html', threshold=pf_threshold)


@app.route('/api/data', methods=['POST'])
def receive_data():
    """Receive 3-phase sensor data from ESP32."""
    try:
        data = request.get_json()
        if not data:
            return jsonify({"error": "No JSON data received"}), 400

        data['timestamp'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Ensure all phase keys exist
        for key in PHASE_KEYS:
            if key not in data:
                data[key] = {"voltage": 0, "current": 0, "power_factor": 0,
                             "real_power": 0, "apparent_power": 0, "reactive_power": 0}

        for field in ['overall_pf', 'total_real_power', 'total_apparent_power', 'total_reactive_power']:
            if field not in data:
                data[field] = 0.0

        readings.append(data)

        # Console output
        opf = data['overall_pf']
        status = "✅" if opf >= pf_threshold else "⚠️ LOW"
        print(f"📡 [{data['timestamp']}] Overall PF={opf:.3f} {status}")
        for i, key in enumerate(PHASE_KEYS):
            p = data[key]
            print(f"   Phase {PHASE_NAMES[i]}: V={p['voltage']:.1f}V "
                  f"I={p['current']:.3f}A PF={p['power_factor']:.3f} "
                  f"P={p['real_power']:.1f}W")

        # Background tasks
        threading.Thread(target=log_to_excel, args=(data.copy(),), daemon=True).start()
        threading.Thread(target=send_notification, args=(data.copy(),), daemon=True).start()

        return jsonify({"status": "ok"}), 200

    except Exception as e:
        print(f"❌ Error: {e}")
        return jsonify({"error": str(e)}), 500


@app.route('/api/readings')
def get_readings():
    count = request.args.get('count', 50, type=int)
    recent = list(readings)[-count:]
    return jsonify({"readings": recent, "threshold": pf_threshold, "total_count": len(readings)})


@app.route('/api/latest')
def get_latest():
    if readings:
        return jsonify({"reading": readings[-1], "threshold": pf_threshold})
    return jsonify({"reading": None, "threshold": pf_threshold})


@app.route('/api/stats')
def get_stats():
    if not readings:
        return jsonify({"error": "No data yet"})

    recent = list(readings)
    result = {"count": len(recent), "threshold": pf_threshold, "phases": {}}

    for i, key in enumerate(PHASE_KEYS):
        pfs = [r[key]['power_factor'] for r in recent if r.get(key, {}).get('power_factor', 0) > 0]
        voltages = [r[key]['voltage'] for r in recent if key in r]
        currents = [r[key]['current'] for r in recent if key in r]

        result["phases"][PHASE_NAMES[i]] = {
            "avg_pf": round(sum(pfs) / len(pfs), 3) if pfs else 0,
            "min_pf": round(min(pfs), 3) if pfs else 0,
            "max_pf": round(max(pfs), 3) if pfs else 0,
            "avg_voltage": round(sum(voltages) / len(voltages), 1) if voltages else 0,
            "avg_current": round(sum(currents) / len(currents), 3) if currents else 0,
            "low_pf_count": sum(1 for p in pfs if p < pf_threshold),
        }

    overall_pfs = [r.get('overall_pf', 0) for r in recent if r.get('overall_pf', 0) > 0]
    result["overall"] = {
        "avg_pf": round(sum(overall_pfs) / len(overall_pfs), 3) if overall_pfs else 0,
        "min_pf": round(min(overall_pfs), 3) if overall_pfs else 0,
        "max_pf": round(max(overall_pfs), 3) if overall_pfs else 0,
    }

    return jsonify(result)


# ════════════════════════════════════════════════════════════
#  SIMULATION MODE — 3-Phase
# ════════════════════════════════════════════════════════════
def simulate_data():
    """Generate fake 3-phase sensor data."""
    import requests as req

    print("🔄 Simulation mode: generating fake 3-phase data...")
    print(f"   Sending to: http://127.0.0.1:{DEFAULT_PORT}/api/data")
    print("   Press Ctrl+C to stop.\n")
    time.sleep(2)

    t = 0
    while True:
        try:
            phases = {}
            total_real = 0
            total_apparent = 0

            for i, key in enumerate(PHASE_KEYS):
                # Each phase has slightly different characteristics
                base_pf = 0.90 + 0.05 * math.sin(t * 0.1 + i * 2.094)
                if random.random() < 0.12:
                    base_pf -= random.uniform(0.1, 0.3)
                pf = max(0.3, min(1.0, base_pf + random.gauss(0, 0.02)))

                voltage = 220 + random.gauss(0, 4) + i * 2  # Slight phase offset
                current = 2.0 + random.gauss(0, 0.3) + i * 0.5
                apparent = voltage * current
                real = apparent * pf
                reactive = math.sqrt(max(0, apparent**2 - real**2))

                phases[key] = {
                    "voltage": round(voltage, 2),
                    "current": round(current, 3),
                    "power_factor": round(pf, 3),
                    "real_power": round(real, 2),
                    "apparent_power": round(apparent, 2),
                    "reactive_power": round(reactive, 2)
                }
                total_real += real
                total_apparent += apparent

            overall_pf = total_real / total_apparent if total_apparent > 0 else 0
            total_reactive = math.sqrt(max(0, total_apparent**2 - total_real**2))

            data = {**phases,
                    "overall_pf": round(abs(overall_pf), 3),
                    "total_real_power": round(total_real, 2),
                    "total_apparent_power": round(total_apparent, 2),
                    "total_reactive_power": round(total_reactive, 2)}

            req.post(f"http://127.0.0.1:{DEFAULT_PORT}/api/data", json=data, timeout=5)
            t += 1
            time.sleep(2)

        except KeyboardInterrupt:
            print("\n🛑 Simulation stopped.")
            break
        except Exception as e:
            print(f"⚠  Simulation error: {e}")
            time.sleep(3)


# ════════════════════════════════════════════════════════════
#  MAIN
# ════════════════════════════════════════════════════════════
def main():
    global pf_threshold, DEFAULT_PORT

    parser = argparse.ArgumentParser(description="3-Phase Power Factor Monitor Server")
    parser.add_argument('--port', type=int, default=5000, help='Server port (default: 5000)')
    parser.add_argument('--threshold', type=float, default=0.85, help='PF threshold (default: 0.85)')
    parser.add_argument('--simulate', action='store_true', help='Run with simulated 3-phase data')
    args = parser.parse_args()

    pf_threshold = args.threshold
    DEFAULT_PORT = args.port

    init_excel()

    print("=" * 60)
    print("  ⚡ 3-PHASE POWER FACTOR MONITOR SERVER")
    print("=" * 60)
    print(f"  📡 Server:       http://0.0.0.0:{args.port}")
    print(f"  🎯 Threshold:    {pf_threshold}")
    print(f"  📊 Excel:        {os.path.abspath(EXCEL_FILE)}")
    print(f"  🌐 Dashboard:    http://localhost:{args.port}")
    print(f"  🔔 Notifications: {'Enabled' if NOTIFICATIONS_AVAILABLE else 'Disabled'}")
    print(f"  📐 Phases:       R, Y, B (3-phase)")
    print("=" * 60)
    print()

    if args.simulate:
        sim_thread = threading.Thread(target=simulate_data, daemon=True)
        sim_thread.start()

    app.run(host='0.0.0.0', port=args.port, debug=False, use_reloader=False)


if __name__ == '__main__':
    main()
